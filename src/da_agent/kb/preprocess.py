"""KB preprocessing pipeline (sync, blocking — runs in an executor thread).

Stages: sheet inventory -> region detection -> header/dtype inference ->
column profiling -> cross-sheet FK inference -> manifest assembly. The
output is a `Manifest` ready for atomic write to `kb/<id>/manifest.json`.

Edge cases (spec §5.2):

- **Huge files**: opened with `read_only=True, data_only=True`. Per-column
  profiling samples up to `PROFILE_ROW_CAP` rows; cardinality reports a
  `cardinality_capped` flag when truncated.
- **Multiple tables per sheet**: region detection clusters non-empty cells,
  splitting at any fully-blank row OR column inside the bounding box.
- **Merged cells / messy headers**: merged ranges are propagated in memory
  before header inference (top-left value fills the rest of the merge).
  Multi-row headers collapse via `f"{r1}_{r2}"` join.
- **Mixed types in a column**: dtype is `mixed` when the value classifier
  yields more than one bucket; numeric/temporal stats only emit when the
  bucket is unambiguously numeric/temporal.
"""

from __future__ import annotations

import time
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from itertools import islice
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .manifest import Column, Manifest, Region, Relationship, SheetSummary

# --------------------------------------------------------------------------- #
# Tunables (kept here so tests can monkey-patch)
# --------------------------------------------------------------------------- #
PROFILE_ROW_CAP = 50_000  # rows scanned per region for column profiling
SAMPLE_ROW_LIMIT = 5  # rows shown in `region.sample_rows`
SAMPLE_VALUE_LIMIT = 5  # distinct values in `column.sample_values`
MIN_REGION_AREA = 2  # drop 1x1 clusters (labels); allow narrow 1-col tables
FK_MIN_CONFIDENCE = 0.8  # |a ∩ b| / |a| threshold to emit a Relationship
HEADER_STRING_RATIO = 0.6  # >=60% of header-row cells must be strings


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #
def build_manifest(raw_path: Path, kb_id: str) -> Manifest:
    """Profile `raw_path` and return a Manifest. Sync — call via `to_thread`."""
    wb = load_workbook(raw_path, read_only=True, data_only=True)
    try:
        # Pre-pass to gather sheet titles so PK heuristic can rule out
        # columns whose stem matches another sheet (those are FK candidates,
        # not PK candidates).
        other_sheet_stems = {ws.title.lower().rstrip("s") for ws in wb.worksheets}
        sheets = [_profile_sheet(ws, other_sheet_stems) for ws in wb.worksheets]
    finally:
        wb.close()
    relationships = _infer_relationships(sheets)
    return Manifest(
        kb_id=kb_id,
        source_filename=raw_path.name,
        generated_at=time.time(),
        sheets=sheets,
        relationships=relationships,
    )


# --------------------------------------------------------------------------- #
# Sheet -> regions
# --------------------------------------------------------------------------- #
@dataclass(slots=True)
class _BBox:
    rmin: int
    cmin: int
    rmax: int
    cmax: int

    @property
    def area(self) -> int:
        return (self.rmax - self.rmin + 1) * (self.cmax - self.cmin + 1)


def _profile_sheet(
    ws: Worksheet, other_sheet_stems: set[str] | None = None
) -> SheetSummary:
    grid = _materialize_sheet(ws)
    occ = _occupancy(grid)
    if not occ:
        return SheetSummary(name=ws.title, dims={"rows": 0, "cols": 0}, regions=[])
    regions: list[Region] = []
    for box in _detect_regions(occ):
        if box.area < MIN_REGION_AREA:
            continue
        regions.append(_profile_region(ws.title, grid, box, other_sheet_stems or set()))
    rows = max((r for r, _ in occ), default=0)
    cols = max((c for _, c in occ), default=0)
    return SheetSummary(
        name=ws.title, dims={"rows": rows, "cols": cols}, regions=regions
    )


def _materialize_sheet(ws: Worksheet) -> list[list[Any]]:
    """Materialize the sheet into a dense 2D list and propagate merged-cell
    values so a multi-cell merge looks like a normal cell to detection."""
    grid: list[list[Any]] = [list(row) for row in ws.iter_rows(values_only=True)]
    if not grid:
        return grid
    # `read_only=True` flattens merged cells into a single value at the
    # top-left and Nones elsewhere. Propagate so header inference sees the
    # value across the whole merged range. `ReadOnlyWorksheet` only exposes
    # `merged_cells` on some openpyxl versions, so guard the attr access.
    mc = getattr(ws, "merged_cells", None)
    merged_ranges = list(getattr(mc, "ranges", []) or [])
    for mr in merged_ranges:
        rmin, cmin, rmax, cmax = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        if rmin - 1 >= len(grid) or cmin - 1 >= len(grid[rmin - 1]):
            continue
        anchor = grid[rmin - 1][cmin - 1]
        if anchor is None:
            continue
        for r in range(rmin, rmax + 1):
            row = grid[r - 1] if r - 1 < len(grid) else []
            for c in range(cmin, cmax + 1):
                if c - 1 < len(row):
                    row[c - 1] = anchor
    return grid


def _occupancy(grid: list[list[Any]]) -> set[tuple[int, int]]:
    """Set of (row_1based, col_1based) cells that are non-blank."""
    occ: set[tuple[int, int]] = set()
    for r, row in enumerate(grid, start=1):
        for c, v in enumerate(row, start=1):
            if not _is_blank(v):
                occ.add((r, c))
    return occ


def _detect_regions(occ: set[tuple[int, int]]) -> list[_BBox]:
    """Cluster occupied cells into rectangular bboxes split at fully-blank
    rows/columns inside the overall bounding rectangle.

    The algorithm: take the global bbox, split it horizontally at every fully
    blank row, recurse into each strip and split vertically at every fully
    blank column. The result is a list of bboxes, each contiguous along both
    axes (no fully blank row/col inside).
    """
    if not occ:
        return []
    rmin = min(r for r, _ in occ)
    rmax = max(r for r, _ in occ)
    cmin = min(c for _, c in occ)
    cmax = max(c for _, c in occ)
    boxes: list[_BBox] = []
    width = cmax - cmin + 1
    # Narrow occupancies (≤1 column wide) intentionally skip the row-axis
    # split: every null in a single-column table is data, not a region
    # separator. Splitting would shred a sparse column into 1-row regions
    # that fail the area filter.
    if width <= 1:
        boxes.append(_BBox(rmin=rmin, cmin=cmin, rmax=rmax, cmax=cmax))
        return boxes
    for r_lo, r_hi in _contiguous_axis(occ, axis=0, lo=rmin, hi=rmax):
        sub = {(r, c) for r, c in occ if r_lo <= r <= r_hi}
        c_lo_local = min(c for _, c in sub)
        c_hi_local = max(c for _, c in sub)
        for c_lo, c_hi in _contiguous_axis(sub, axis=1, lo=c_lo_local, hi=c_hi_local):
            box = _shrink_to_data(sub, r_lo, c_lo, r_hi, c_hi)
            if box is not None:
                boxes.append(box)
    return boxes


def _contiguous_axis(
    occ: set[tuple[int, int]], *, axis: int, lo: int, hi: int
) -> list[tuple[int, int]]:
    """Yield (start, end) ranges along `axis` (0=row, 1=col) covering only
    non-blank lines. Empty lines act as separators."""
    has_data = {p[axis] for p in occ}
    spans: list[tuple[int, int]] = []
    cur_start: int | None = None
    for i in range(lo, hi + 1):
        if i in has_data:
            if cur_start is None:
                cur_start = i
            cur_end = i
        else:
            if cur_start is not None:
                spans.append((cur_start, cur_end))
                cur_start = None
    if cur_start is not None:
        spans.append((cur_start, cur_end))
    return spans


def _shrink_to_data(
    occ: set[tuple[int, int]], r_lo: int, c_lo: int, r_hi: int, c_hi: int
) -> _BBox | None:
    sub = {(r, c) for r, c in occ if r_lo <= r <= r_hi and c_lo <= c <= c_hi}
    if not sub:
        return None
    return _BBox(
        rmin=min(r for r, _ in sub),
        cmin=min(c for _, c in sub),
        rmax=max(r for r, _ in sub),
        cmax=max(c for _, c in sub),
    )


# --------------------------------------------------------------------------- #
# Region -> Region(profiled)
# --------------------------------------------------------------------------- #
def _profile_region(
    sheet_name: str,
    grid: list[list[Any]],
    box: _BBox,
    other_sheet_stems: set[str],
) -> Region:
    rows = [
        [
            grid[r - 1][c - 1] if c - 1 < len(grid[r - 1]) else None
            for c in range(box.cmin, box.cmax + 1)
        ]
        for r in range(box.rmin, box.rmax + 1)
        if r - 1 < len(grid)
    ]
    header_local_idx, low_conf = _infer_header_row(rows)
    header_row_abs = box.rmin + header_local_idx
    headers = _collapse_multirow_headers(rows[: header_local_idx + 1])
    n_cols = len(headers)
    data_rows = rows[header_local_idx + 1 : header_local_idx + 1 + PROFILE_ROW_CAP]
    columns: list[Column] = []
    for ci in range(n_cols):
        values = [r[ci] if ci < len(r) else None for r in data_rows]
        columns.append(_profile_column(headers[ci], values))
    sample_rows = [
        [r[ci] if ci < len(r) else None for ci in range(n_cols)]
        for r in data_rows[:SAMPLE_ROW_LIMIT]
    ]
    region_id = f"{sheet_name}!{get_column_letter(box.cmin)}{box.rmin}"
    range_str = (
        f"{get_column_letter(box.cmin)}{box.rmin}"
        f":{get_column_letter(box.cmax)}{box.rmax}"
    )
    # Mark PK candidates: an "id-like" column with no nulls, full uniqueness,
    # and an integer-or-string dtype. The id-like name guard keeps generic
    # numeric columns (e.g. `customer_id` in a fact table) from being
    # mis-typed as PKs purely on uniqueness; the FK pass below relies on
    # this distinction.
    n_data = len(data_rows)
    if n_data > 0:
        for col in columns:
            if (
                col.role is None
                and col.null_pct == 0.0
                and col.cardinality == n_data
                and not col.cardinality_capped
                and col.dtype in {"int", "str"}
                and _looks_like_pk(col.name, sheet_name, other_sheet_stems)
            ):
                col.role = "pk?"
    return Region(
        region_id=region_id,
        range=range_str,
        header_row=header_row_abs,
        columns=columns,
        sample_rows=[_to_jsonable(row) for row in sample_rows],
        low_confidence=low_conf,
    )


def _infer_header_row(rows: list[list[Any]]) -> tuple[int, bool]:
    """Pick the first row whose non-null cells are >=`HEADER_STRING_RATIO`
    strings AND no cell is purely numeric. Otherwise fall back to row 0
    with `low_confidence=True`."""
    for i, row in enumerate(rows[:5]):  # only inspect the first 5 rows
        non_null = [v for v in row if not _is_blank(v)]
        if not non_null:
            continue
        n_str = sum(1 for v in non_null if isinstance(v, str))
        n_num = sum(
            1
            for v in non_null
            if isinstance(v, (int, float)) and not isinstance(v, bool)
        )
        if n_num > 0:
            continue
        if n_str / len(non_null) >= HEADER_STRING_RATIO:
            return i, False
    return 0, True


def _collapse_multirow_headers(header_rows: list[list[Any]]) -> list[str]:
    """Join non-blank header cells across multiple rows with `_`."""
    if not header_rows:
        return []
    width = max(len(r) for r in header_rows)
    out: list[str] = []
    for c in range(width):
        parts = []
        for r in header_rows:
            v = r[c] if c < len(r) else None
            if not _is_blank(v):
                parts.append(str(v).strip())
        out.append("_".join(parts) if parts else f"col_{c + 1}")
    return out


# --------------------------------------------------------------------------- #
# Column profiling
# --------------------------------------------------------------------------- #
def _profile_column(name: str, values: list[Any]) -> Column:
    non_null = [v for v in values if not _is_blank(v)]
    n_total = len(values)
    null_pct = 0.0 if n_total == 0 else round(1 - len(non_null) / n_total, 4)

    buckets = Counter(_classify(v) for v in non_null)
    if not buckets:
        return Column(name=name, dtype="str", cardinality=0, null_pct=null_pct)
    if len(buckets) == 1:
        dtype = next(iter(buckets))
    else:
        # int + float can collapse into "float"; everything else is "mixed".
        keys = set(buckets)
        if keys <= {"int", "float"}:
            dtype = "float"
        elif keys <= {"date", "datetime"}:
            dtype = "datetime"
        else:
            dtype = "mixed"

    distinct: set[Any] = set()
    capped = False
    for v in non_null:
        distinct.add(v)
        if len(distinct) > PROFILE_ROW_CAP:
            capped = True
            break
    cardinality = len(distinct)

    col = Column(
        name=name,
        dtype=dtype,
        cardinality=cardinality,
        null_pct=null_pct,
        cardinality_capped=capped,
    )
    if dtype in {"int", "float"}:
        nums = [
            v
            for v in non_null
            if isinstance(v, (int, float)) and not isinstance(v, bool)
        ]
        if nums:
            col.min = min(nums)
            col.max = max(nums)
    elif dtype in {"date", "datetime"}:
        temps = [v for v in non_null if isinstance(v, (date, datetime))]
        if temps:
            col.min = str(min(temps))
            col.max = str(max(temps))
    col.sample_values = _to_jsonable(list(islice(distinct, SAMPLE_VALUE_LIMIT)))
    return col


def _looks_like_pk(col_name: str, sheet_name: str, other_sheet_stems: set[str]) -> bool:
    """Heuristic: id-like column whose name does NOT reference another
    sheet's entity. So `id` and `order_id` in `Sales` are PK candidates,
    but `customer_id` in `Sales` (referencing the `Customers` sheet) is
    NOT -- it's an FK. `other_sheet_stems` is the set of *other* sheet
    titles in singular form (lowercased, trailing `s` stripped)."""
    name = col_name.lower()
    if name == "id":
        return True
    sheet_lc = sheet_name.lower()
    sheet_singular = sheet_lc.rstrip("s")
    # Direct sheet-self matches: `<sheet>_id`, `<singular>_id`, `<sheet>id`.
    if name in {f"{sheet_lc}_id", f"{sheet_singular}_id", f"{sheet_lc}id"}:
        return True
    # Generic `<noun>_id` pattern, accepted iff the noun is not another
    # sheet's entity. Excludes own sheet stems too (those already matched).
    if name.endswith("_id"):
        stem = name[:-3]
        if stem and stem not in other_sheet_stems - {sheet_singular}:
            return True
    return False


def _classify(v: Any) -> str:
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float"
    if isinstance(v, datetime):
        return "datetime"
    if isinstance(v, date):
        return "date"
    return "str"


# --------------------------------------------------------------------------- #
# Cross-sheet FK inference
# --------------------------------------------------------------------------- #
def _infer_relationships(sheets: list[SheetSummary]) -> list[Relationship]:
    """Emit Relationship rows for FK candidates with high overlap.

    Heuristic: source.col looks like an FK iff (a) some other sheet has a
    column whose role is "pk?", (b) the dtypes match (or both numeric), and
    (c) the source's sample values are a subset of the target's sample
    values at >= FK_MIN_CONFIDENCE ratio. We intentionally use sample_values
    rather than the full column -- streaming overlap on raw rows would mean
    re-reading the workbook, defeating the manifest's purpose.
    """
    pks: list[tuple[str, str, Column]] = []  # (sheet_name, region_id, column)
    for sheet in sheets:
        for region in sheet.regions:
            for col in region.columns:
                if col.role == "pk?":
                    pks.append((sheet.name, region.region_id, col))

    rels: list[Relationship] = []
    for sheet in sheets:
        for region in sheet.regions:
            for col in region.columns:
                if col.role == "pk?":
                    continue
                target = _best_pk_match(sheet.name, col, pks)
                if target is None:
                    continue
                target_sheet, _, target_col = target
                target_set = set(target_col.sample_values)
                if not target_set:
                    continue
                src = [v for v in col.sample_values if v in target_set]
                ratio = len(src) / max(len(col.sample_values), 1)
                if ratio < FK_MIN_CONFIDENCE:
                    continue
                col.role = f"fk?->{target_sheet}.{target_col.name}"
                rels.append(
                    Relationship(
                        from_=f"{sheet.name}.{col.name}",
                        to=f"{target_sheet}.{target_col.name}",
                        confidence=round(ratio, 2),
                    )
                )
    return rels


def _best_pk_match(
    src_sheet: str, src_col: Column, pks: list[tuple[str, str, Column]]
) -> tuple[str, str, Column] | None:
    """Pick a PK candidate whose name relates to the source column. Prefer
    suffix `_id` matches (`customer_id` -> `Customers.id`); fall back to
    exact name match across sheets."""
    suffix = src_col.name.lower().rsplit("_", 1)[-1] if "_" in src_col.name else None
    stem = src_col.name.lower().removesuffix("_id")
    for pk_sheet, region_id, pk_col in pks:
        if pk_sheet == src_sheet:
            continue
        if not _dtypes_compatible(src_col.dtype, pk_col.dtype):
            continue
        pk_name_lc = pk_col.name.lower()
        sheet_lc = pk_sheet.lower().rstrip("s")
        if (
            (suffix == pk_name_lc)
            or (stem and (stem == sheet_lc or stem == pk_sheet.lower()))
            or (src_col.name.lower() == pk_name_lc)
        ):
            return (pk_sheet, region_id, pk_col)
    return None


def _dtypes_compatible(a: str, b: str) -> bool:
    if a == b:
        return True
    numeric = {"int", "float"}
    return a in numeric and b in numeric


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _to_jsonable(values: Iterable[Any]) -> list[Any]:
    """Convert datetime/date to ISO strings; passthrough otherwise."""
    out: list[Any] = []
    for v in values:
        if isinstance(v, (datetime, date)):
            out.append(v.isoformat())
        else:
            out.append(v)
    return out
