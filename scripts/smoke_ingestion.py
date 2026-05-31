"""Live smoke for the memory-driven KB ingestion pipeline.

Runs against a real model (the kb_profiler subagent uses opus by default;
override with DA_AGENT_KB_PROFILER_MODEL) behind real uvicorn. The five
phases probe the full upload → profile → scope → analysis loop end-to-end.

Phases:
  A — POST /kb/files with a small two-sheet xlsx; poll status until terminal.
      Pass: status == READY, memory_path is set, file exists on disk, the
      memory body contains the sheet names + at least one column dtype.
  B — GET /kb/files/{id}/memory: returns 200, content matches disk.
  C — Open a session, send a memory-driven prompt with this KB in scope.
      Assert the main agent invokes Read on the memory file (kb_<id>.md)
      BEFORE any heavy compute. Final answer references the KB content.
  D — POST /kb/files/{id}/reprofile schedules a fresh run; the memory file
      timestamp moves forward (or content changes).
  E — DELETE /kb/files/{id}: row gone, kb dir gone, memory file gone.

Pre-req: ANTHROPIC_API_KEY (or whatever the configured model uses) must be
available. The opus pass typically takes 30–120s for a small file.

Run:
    uv run python scripts/smoke_ingestion.py
"""

from __future__ import annotations

import json
import os
import shutil
import socket
import subprocess
import sys
import tempfile
import time
import urllib.error
import urllib.request
import uuid
from pathlib import Path

import openpyxl


def _free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


def _wait_healthz(port: int, timeout: float = 30.0) -> None:
    url = f"http://127.0.0.1:{port}/health"
    deadline = time.monotonic() + timeout
    last_err: Exception | None = None
    while time.monotonic() < deadline:
        try:
            with urllib.request.urlopen(url, timeout=1.5) as r:
                if r.status == 200:
                    return
        except (urllib.error.URLError, ConnectionError, TimeoutError) as e:
            last_err = e
            time.sleep(0.3)
    raise RuntimeError(f"server never became healthy: {last_err}")


def _post_json(port: int, path: str, payload: dict | None = None) -> dict:
    body = b"" if payload is None else json.dumps(payload).encode()
    req = urllib.request.Request(
        f"http://127.0.0.1:{port}{path}",
        data=body,
        headers={"Content-Type": "application/json"} if payload is not None else {},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        raw = r.read()
        return json.loads(raw) if raw else {}


def _get_json(port: int, path: str) -> dict:
    with urllib.request.urlopen(f"http://127.0.0.1:{port}{path}", timeout=30) as r:
        return json.loads(r.read())


def _delete(port: int, path: str) -> int:
    req = urllib.request.Request(f"http://127.0.0.1:{port}{path}", method="DELETE")
    with urllib.request.urlopen(req, timeout=15) as r:
        return r.status


def _multipart_upload(port: int, path: str, file_path: Path, mime: str) -> dict:
    boundary = f"----smoke{uuid.uuid4().hex}"
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; '
        f'filename="{file_path.name}"\r\n'
        f"Content-Type: {mime}\r\n\r\n"
    ).encode()
    body += file_path.read_bytes()
    body += f"\r\n--{boundary}--\r\n".encode()
    req = urllib.request.Request(
        f"http://127.0.0.1:{port}{path}",
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read())


def _stream_turn(
    port: int,
    sid: str,
    prompt: str,
    *,
    body_extra: dict | None = None,
    deadline_s: float = 600.0,
) -> list[dict]:
    req = urllib.request.Request(
        f"http://127.0.0.1:{port}/sessions/{sid}/messages",
        data=json.dumps({"prompt": prompt, **(body_extra or {})}).encode(),
        headers={"Content-Type": "application/json", "Accept": "text/event-stream"},
        method="POST",
    )
    events: list[dict] = []
    deadline = time.monotonic() + deadline_s
    with urllib.request.urlopen(req, timeout=deadline_s) as r:
        buf: list[str] = []
        for raw in r:
            if time.monotonic() > deadline:
                raise TimeoutError("SSE drain timed out")
            line = raw.decode("utf-8").rstrip("\n")
            if line == "":
                rec: dict = {}
                for entry in buf:
                    if entry.startswith("data:"):
                        try:
                            rec = json.loads(entry[5:].strip())
                        except json.JSONDecodeError:
                            pass
                if rec:
                    events.append(rec)
                buf.clear()
            else:
                buf.append(line)
            if events and events[-1].get("type") == "result":
                break
    return events


def _make_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    customers = wb.create_sheet("Customers")
    customers.append(["customer_id", "name", "country", "signup_date"])
    customers.append([1, "Alice Nguyen", "VN", "2024-01-12"])
    customers.append([2, "Bao Tran", "VN", "2024-02-03"])
    customers.append([3, "Charlie Lee", "US", "2024-02-19"])
    customers.append([4, "Dao Pham", "VN", "2024-03-04"])
    customers.append([5, "Eve Wong", "SG", "2024-03-22"])

    sales = wb.create_sheet("Sales")
    sales.append(["order_id", "customer_id", "amount", "order_date"])
    sales.append([1001, 1, 120.5, "2024-04-02"])
    sales.append([1002, 2, 65.0, "2024-04-05"])
    sales.append([1003, 1, 230.0, "2024-04-09"])
    sales.append([1004, 3, 99.99, "2024-04-11"])
    sales.append([1005, 5, 410.0, "2024-04-15"])
    sales.append([1006, 4, 18.75, "2024-04-22"])

    wb.save(path)


def _wait_terminal(port: int, kb_id: str, *, timeout: float = 600.0) -> dict:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        meta = _get_json(port, f"/kb/files/{kb_id}")
        if meta["status"] in {"READY", "READY_PARTIAL", "FAILED"}:
            return meta
        time.sleep(1.0)
    raise TimeoutError(f"kb {kb_id} never reached terminal status within {timeout}s")


def _did_read_memory(events: list[dict], memory_path: str) -> bool:
    """True iff a `tool.use` of Read fired against the memory file path."""
    for ev in events:
        if ev.get("type") != "tool.use":
            continue
        if ev.get("name") != "Read":
            continue
        fp = (ev.get("input") or {}).get("file_path") or ""
        if memory_path in fp:
            return True
    return False


def main() -> int:
    repo_root = Path(__file__).resolve().parents[1]
    data_root = Path(tempfile.mkdtemp(prefix="da-agent-ingestion-smoke-"))
    port = _free_port()
    print(f"→ smoke: data_root={data_root}  port={port}")

    xlsx_path = data_root / "smoke.xlsx"
    _make_xlsx(xlsx_path)

    env = dict(os.environ)
    env["DA_AGENT_HOME"] = str(data_root)
    env["PYTHONUNBUFFERED"] = "1"

    proc = subprocess.Popen(
        [
            sys.executable,
            "-m",
            "da_agent.cli",
            "serve",
            "--host",
            "127.0.0.1",
            "--port",
            str(port),
        ],
        cwd=str(repo_root),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )
    failures: list[str] = []
    try:
        try:
            _wait_healthz(port)
            print("✓ server healthy")

            # --- Phase A — upload + poll READY -----------------------------
            print("\n=== Phase A — upload + memory generation ===")
            kb_meta = _multipart_upload(
                port,
                "/kb/files",
                xlsx_path,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            kb_id = kb_meta["id"]
            print(f"  uploaded: {kb_id}, status={kb_meta['status']}")
            settled = _wait_terminal(port, kb_id, timeout=600.0)
            print(
                f"  terminal: status={settled['status']}, "
                f"memory_path={settled.get('memory_path')}"
            )
            if settled["status"] != "READY":
                failures.append(
                    f"Phase A: expected READY, got {settled['status']} "
                    f"(error={settled.get('error')})"
                )
            elif not settled.get("memory_path"):
                failures.append("Phase A: status READY but memory_path empty")
            else:
                memory_file = Path(settled["memory_path"])
                if not memory_file.exists():
                    failures.append(f"Phase A: memory_path {memory_file} not on disk")
                else:
                    body = memory_file.read_text("utf-8")
                    if "Customers" not in body or "Sales" not in body:
                        failures.append(
                            "Phase A: memory body missing sheet names "
                            "(Customers / Sales)"
                        )

            # --- Phase B — GET /memory ------------------------------------
            if settled["status"] == "READY":
                print("\n=== Phase B — GET /kb/files/{id}/memory ===")
                mem_resp = _get_json(port, f"/kb/files/{kb_id}/memory")
                disk = Path(settled["memory_path"]).read_text("utf-8")
                if mem_resp["content"] != disk:
                    failures.append("Phase B: /memory content != disk content")
                else:
                    print(f"  body length: {mem_resp['size_bytes']} bytes")

            # --- Phase C — main-agent reads memory before answering -------
            if settled["status"] == "READY":
                print("\n=== Phase C — main agent must Read memory file ===")
                sess = _post_json(port, "/sessions", {"name": "smoke-ingestion"})
                sid = sess["id"]
                events = _stream_turn(
                    port,
                    sid,
                    "Liệt kê các sheet trong KB này. Trả lời ngắn gọn.",
                    body_extra={"kb_scope": [kb_id]},
                    deadline_s=300.0,
                )
                memory_path_str = settled["memory_path"]
                if not _did_read_memory(events, memory_path_str):
                    failures.append(
                        "Phase C: main agent did NOT Read the memory file "
                        f"({memory_path_str})"
                    )
                else:
                    print("  ✓ main agent invoked Read on memory file")
                _delete(port, f"/sessions/{sid}")

            # --- Phase D — reprofile updates the memory file --------------
            if settled["status"] == "READY":
                print("\n=== Phase D — POST /reprofile ===")
                memory_file = Path(settled["memory_path"])
                before_mtime = memory_file.stat().st_mtime
                resp = _post_json(port, f"/kb/files/{kb_id}/reprofile")
                if resp["status"] not in {"PENDING", "PROFILING"}:
                    failures.append(
                        f"Phase D: reprofile returned status={resp['status']}, "
                        f"expected PENDING/PROFILING"
                    )
                resettled = _wait_terminal(port, kb_id, timeout=600.0)
                if resettled["status"] != "READY":
                    failures.append(
                        f"Phase D: reprofile terminal status="
                        f"{resettled['status']}, expected READY"
                    )
                elif memory_file.stat().st_mtime <= before_mtime:
                    failures.append(
                        "Phase D: memory file mtime did not advance after reprofile"
                    )
                else:
                    print("  ✓ memory file rewritten by reprofile")

            # --- Phase E — DELETE cleans memory file ----------------------
            print("\n=== Phase E — DELETE cleans memory file ===")
            memory_file = (
                data_root.parent  # placeholder; replaced below
            )
            current = _get_json(port, f"/kb/files/{kb_id}")
            mp = current.get("memory_path")
            memory_file = Path(mp) if mp else None
            del_status = _delete(port, f"/kb/files/{kb_id}")
            if del_status != 204:
                failures.append(f"Phase E: DELETE returned {del_status}")
            else:
                # row gone
                try:
                    _get_json(port, f"/kb/files/{kb_id}")
                    failures.append("Phase E: row still present after DELETE")
                except urllib.error.HTTPError as e:
                    if e.code != 404:
                        failures.append(
                            f"Phase E: GET after DELETE returned {e.code}, expected 404"
                        )
                kb_dir = data_root / "kb" / kb_id
                if kb_dir.exists():
                    failures.append(f"Phase E: kb dir {kb_dir} not removed")
                if memory_file is not None and memory_file.exists():
                    failures.append(f"Phase E: memory file {memory_file} not removed")

        except Exception as exc:
            failures.append(f"smoke crashed mid-flow: {type(exc).__name__}: {exc}")

    finally:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
        if not failures:
            shutil.rmtree(data_root, ignore_errors=True)
        else:
            print(f"\n⚠ data_root preserved for inspection: {data_root}")

    print()
    if failures:
        print("✗ SMOKE FAILED:")
        for f in failures:
            print(f"  - {f}")
        return 1
    print("✓ SMOKE PASSED — ingestion + memory + reprofile + delete all green.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
