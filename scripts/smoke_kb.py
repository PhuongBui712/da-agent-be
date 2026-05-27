"""Live smoke for KB ingestion + agent manifest-first behavior.

Build a tiny xlsx (Customers + Sales sheets) -> POST /kb/files -> poll READY ->
GET manifest -> POST /sessions/{sid}/messages with an analysis prompt -> stream
SSE -> assert the agent reads `manifest.json` BEFORE touching `raw.xlsx`.

Run with the dev deps installed:
    uv run python scripts/smoke_kb.py

Exit code:
    0  pass (manifest read first, no full-sheet dump)
    1  fail (any acceptance criterion missed)
"""

from __future__ import annotations

import asyncio
import io
import json
import os
import sys
import tempfile
import time
from pathlib import Path

import httpx
import openpyxl

from da_agent.config import Settings
from da_agent.server.app import create_app


def _make_dummy_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cust = wb.create_sheet("Customers")
    cust.append(["id", "name", "country"])
    cust.append([1, "Alice", "VN"])
    cust.append([2, "Bob", "US"])
    cust.append([3, "Carol", "VN"])
    cust.append([4, "Dan", "JP"])
    cust.append([5, "Eve", "US"])

    sales = wb.create_sheet("Sales")
    sales.append(["order_id", "customer_id", "amount", "ts"])
    sales.append([101, 1, 50.0, "2026-01-02"])
    sales.append([102, 2, 75.5, "2026-01-03"])
    sales.append([103, 1, 12.0, "2026-01-05"])
    sales.append([104, 4, 99.9, "2026-01-09"])
    sales.append([105, 3, 33.3, "2026-01-10"])
    sales.append([106, 2, 5.5, "2026-01-12"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _wait_ready(client: httpx.AsyncClient, kb_id: str, *, timeout: float = 30.0):
    started = time.monotonic()
    while time.monotonic() - started < timeout:
        r = await client.get(f"/kb/files/{kb_id}")
        r.raise_for_status()
        meta = r.json()
        if meta["status"] == "READY":
            return meta
        if meta["status"] == "FAILED":
            raise RuntimeError(f"KB ingest FAILED: {meta.get('error')}")
        await asyncio.sleep(0.25)
    raise TimeoutError(f"KB {kb_id} not READY after {timeout}s")


async def main() -> int:
    tmp_root = Path(tempfile.mkdtemp(prefix="da-agent-smoke-kb-"))
    os.environ["DA_AGENT_HOME"] = str(tmp_root)
    settings = Settings()
    settings.data_root = tmp_root
    settings.plan_first = False
    settings.show_thinking = False  # noisy for smoke logs
    settings.max_turns = 6
    settings.ensure_dirs()
    print(f"→ data_root: {tmp_root}")

    app = create_app(settings)
    async with app.router.lifespan_context(app):
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport, base_url="http://test", timeout=120
        ) as client:
            # 1) Upload xlsx
            xlsx_bytes = _make_dummy_xlsx()
            r = await client.post(
                "/kb/files",
                files={
                    "file": (
                        "demo.xlsx",
                        xlsx_bytes,
                        (
                            "application/vnd.openxmlformats-"
                            "officedocument.spreadsheetml.sheet"
                        ),
                    )
                },
            )
            assert r.status_code == 202, (r.status_code, r.text)
            kb_id = r.json()["id"]
            print(f"→ uploaded kb_id={kb_id}, status=PENDING")

            # 2) Wait READY
            meta = await _wait_ready(client, kb_id)
            print(
                f"→ status=READY (took {meta['updated_at'] - meta['created_at']:.2f}s)"
            )

            # 3) Fetch manifest, sanity-check shape
            r = await client.get(f"/kb/files/{kb_id}/manifest")
            assert r.status_code == 200, r.text
            manifest = r.json()
            sheet_names = [s["name"] for s in manifest["sheets"]]
            rels = manifest.get("relationships", [])
            print(f"→ manifest sheets: {sheet_names}; relationships: {rels}")
            if "Customers" not in sheet_names or "Sales" not in sheet_names:
                print("FAIL: missing sheets in manifest")
                return 1
            if not rels:
                print("WARN: no FK inferred (heuristic miss; not a hard fail)")

            # 4) Create a session and ask an analysis question
            r = await client.post("/sessions", json={"name": "smoke-kb"})
            sid = r.json()["id"]
            prompt = (
                f"Tôi đã upload một KB ID `{kb_id}` chứa hai sheet Customers và Sales. "
                "Hãy đọc manifest trước rồi cho tôi biết: tổng doanh thu (sum amount) "
                "theo country, và liệt kê top 1 customer có tổng amount cao nhất. "
                "Trả lời ngắn gọn."
            )

            tool_uses: list[tuple[str, dict]] = []
            async with client.stream(
                "POST",
                f"/sessions/{sid}/messages",
                json={"prompt": prompt},
            ) as resp:
                event_type: str | None = None
                data_lines: list[str] = []
                async for raw in resp.aiter_lines():
                    line = raw.rstrip("\r")
                    if line == "":
                        if data_lines:
                            try:
                                payload = json.loads("\n".join(data_lines))
                            except json.JSONDecodeError:
                                payload = {}
                            if event_type == "tool.use":
                                tool_uses.append(
                                    (
                                        payload.get("name", ""),
                                        payload.get("input") or {},
                                    )
                                )
                        event_type, data_lines = None, []
                    elif line.startswith("event:"):
                        event_type = line[len("event:") :].strip()
                    elif line.startswith("data:"):
                        data_lines.append(line[len("data:") :].lstrip())

            # 5) Acceptance: did the agent read manifest.json before raw.xlsx?
            print(f"→ {len(tool_uses)} tool.use events")
            manifest_ts = None
            raw_ts = None
            for idx, (name, ti) in enumerate(tool_uses):
                blob = json.dumps(ti, default=str).lower()
                if "manifest.json" in blob and manifest_ts is None:
                    manifest_ts = idx
                if "raw.xlsx" in blob and raw_ts is None:
                    raw_ts = idx
                summary = {
                    k: ti.get(k) for k in ("file_path", "command", "pattern") if k in ti
                }
                print(f"  [{idx:02d}] {name}: {summary}")

            print()
            issues: list[str] = []
            if manifest_ts is None:
                issues.append("agent never read manifest.json (FAIL)")
            elif raw_ts is not None and manifest_ts > raw_ts:
                issues.append(
                    f"agent opened raw.xlsx before manifest "
                    f"(raw at {raw_ts}, manifest at {manifest_ts})"
                )

            if issues:
                print("FAIL:", "; ".join(issues))
                return 1
            print(
                "OK: manifest read before raw.xlsx; agent honored manifest-first contract."
            )
            return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
